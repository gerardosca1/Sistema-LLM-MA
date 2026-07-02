import openpyxl
import sys

path = r'C:\Users\luis.acosta\Documents\ProyectosGlobant\G4G\2026\4950- [RAS]- For due date contracts block revenue actions\GLB951-66593-GLB951-66595 - Create a New Project with a Waiver\Test_Cases_GLB951-66593_GLB951-66595.xlsx'

try:
    wb = openpyxl.load_workbook(path)
    print('Sheets:', wb.sheetnames)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f'\n=== Sheet: {sheet} ===')
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 100:
                print(f'Row {i+1}:', row)
except Exception as e:
    print(f'Error: {e}', file=sys.stderr)
