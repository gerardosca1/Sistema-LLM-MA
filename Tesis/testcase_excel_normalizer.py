import argparse
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl


HEADER_ALIASES: Dict[str, List[str]] = {
    "id": ["id", "tc", "test case id", "testcase id", "case id"],
    "summary": ["summary", "title", "test case", "testcase", "name"],
    "objective": ["objective", "objetivo"],
    "preconditions": ["preconditions", "precondition", "pre-condiciones", "precondiciones"],
    "test_data": ["test data", "data", "datos", "datos de prueba"],
    "steps": ["steps", "test steps", "pasos", "procedure", "procedimiento"],
    "expected": ["expected", "expected result", "result", "resultado esperado"],
    "postconditions": ["postconditions", "postcondition", "post-condiciones", "postcondiciones"],
    "priority": ["priority", "prioridad"],
    "test_type": ["test type", "type", "tipo"],
}


def normalize_header(value: Optional[str]) -> str:
    if value is None:
        return ""
    value = str(value).strip().lower()
    value = re.sub(r"\s+", " ", value)
    return value


def detect_header_row(ws, max_scan_rows: int = 30) -> int:
    best_row = 1
    best_score = -1

    for row_idx in range(1, min(max_scan_rows, ws.max_row) + 1):
        cells = [normalize_header(c.value) for c in ws[row_idx]]
        if not any(cells):
            continue

        score = 0
        for cell in cells:
            for aliases in HEADER_ALIASES.values():
                if cell in aliases:
                    score += 1
                    break

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def find_column_map(headers: List[str]) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for idx, raw in enumerate(headers):
        h = normalize_header(raw)
        if not h:
            continue

        for canonical, aliases in HEADER_ALIASES.items():
            if h in aliases and canonical not in col_map:
                col_map[canonical] = idx
                break

    return col_map


def get_cell(row_values: Tuple, col_map: Dict[str, int], key: str) -> str:
    idx = col_map.get(key)
    if idx is None or idx >= len(row_values):
        return ""
    value = row_values[idx]
    return "" if value is None else str(value).strip()


def split_steps(steps_text: str, expected_text: str) -> List[Dict[str, str]]:
    # If expected is already provided, keep one-step structure.
    if steps_text and expected_text:
        return [{"step": "1", "action": steps_text, "expected": expected_text}]

    if not steps_text:
        return []

    # Try to parse multiline numbered steps: "1. action | expected"
    lines = [l.strip() for l in steps_text.splitlines() if l.strip()]
    parsed: List[Dict[str, str]] = []
    for i, line in enumerate(lines, start=1):
        parts = [p.strip() for p in line.split("|")]
        if len(parts) == 2:
            action, expected = parts
        else:
            action, expected = line, ""
        parsed.append({"step": str(i), "action": action, "expected": expected})

    return parsed


def build_jira_description(tc: Dict) -> str:
    lines: List[str] = []

    if tc.get("objective"):
        lines.extend(["### Objective", "", tc["objective"], ""])

    if tc.get("preconditions"):
        lines.extend(["### Preconditions", "", tc["preconditions"], ""])

    if tc.get("test_data"):
        lines.extend(["### Test Data", "", tc["test_data"], ""])

    lines.extend(["### Test Steps", "", "| Step # | Action | Expected Result |", "| --- | --- | --- |"])
    steps = tc.get("steps", [])
    if steps:
        for step in steps:
            lines.append(
                f"| {step.get('step', '')} | {step.get('action', '').replace('|', '/')} | {step.get('expected', '').replace('|', '/')} |"
            )
    else:
        lines.append("| 1 | N/A | N/A |")

    lines.append("")

    if tc.get("postconditions"):
        lines.extend(["### Postconditions", "", tc["postconditions"], ""])

    if tc.get("test_type"):
        lines.extend(["### Test Type", "", tc["test_type"], ""])

    if tc.get("priority"):
        lines.extend(["### Priority", "", tc["priority"], ""])

    return "\n".join(lines).strip()


def normalize_sheet(ws, header_row: Optional[int] = None) -> List[Dict]:
    if header_row is None:
        header_row = detect_header_row(ws)

    headers = ["" if c.value is None else str(c.value) for c in ws[header_row]]
    col_map = find_column_map(headers)

    testcases: List[Dict] = []
    empty_streak = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_values = tuple(c.value for c in ws[row_idx])

        raw_id = get_cell(row_values, col_map, "id")
        raw_summary = get_cell(row_values, col_map, "summary")

        if not raw_id and not raw_summary and not any(v is not None and str(v).strip() for v in row_values):
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue

        empty_streak = 0

        if not raw_summary:
            # Skip non-data lines where we cannot infer summary.
            continue

        tc = {
            "id": raw_id,
            "summary": raw_summary,
            "objective": get_cell(row_values, col_map, "objective"),
            "preconditions": get_cell(row_values, col_map, "preconditions"),
            "test_data": get_cell(row_values, col_map, "test_data"),
            "postconditions": get_cell(row_values, col_map, "postconditions"),
            "priority": get_cell(row_values, col_map, "priority"),
            "test_type": get_cell(row_values, col_map, "test_type"),
        }
        steps_text = get_cell(row_values, col_map, "steps")
        expected_text = get_cell(row_values, col_map, "expected")
        tc["steps"] = split_steps(steps_text, expected_text)
        tc["jira_description"] = build_jira_description(tc)

        testcases.append(tc)

    return testcases


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Normalize test cases from Excel into JSON + markdown summary for faster Jira/Xray loading."
    )
    parser.add_argument("--input", required=True, help="Path to input .xlsx file")
    parser.add_argument("--sheet", help="Sheet name (default: first sheet)")
    parser.add_argument("--header-row", type=int, help="1-based header row. If omitted, auto-detected")
    parser.add_argument("--output-json", default="testcases_normalized.json", help="Output JSON file path")
    parser.add_argument("--output-md", default="testcases_preview.md", help="Output markdown preview path")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    wb = openpyxl.load_workbook(input_path)
    if args.sheet:
        if args.sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
        ws = wb[args.sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    testcases = normalize_sheet(ws, args.header_row)

    out_json = Path(args.output_json)
    out_json.write_text(json.dumps(testcases, ensure_ascii=False, indent=2), encoding="utf-8")

    md_lines = [
        f"# Test Cases Preview ({len(testcases)})",
        "",
        f"Source: `{input_path}`",
        f"Sheet: `{ws.title}`",
        "",
    ]
    for i, tc in enumerate(testcases, start=1):
        title = tc.get("summary", "(No summary)")
        tc_id = tc.get("id", "")
        md_lines.append(f"## {i}. {tc_id} - {title}" if tc_id else f"## {i}. {title}")
        md_lines.append("")
        md_lines.append(tc["jira_description"])
        md_lines.append("\n---\n")

    out_md = Path(args.output_md)
    out_md.write_text("\n".join(md_lines), encoding="utf-8")

    print(f"✅ Parsed test cases: {len(testcases)}")
    print(f"✅ JSON output: {out_json.resolve()}")
    print(f"✅ Markdown preview: {out_md.resolve()}")
    if not testcases:
        print("⚠️  No test cases parsed. Try setting --header-row manually.")


if __name__ == "__main__":
    main()
