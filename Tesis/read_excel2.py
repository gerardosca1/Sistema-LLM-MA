import openpyxl
import sys

path = r'C:\Users\luis.acosta\Documents\ProyectosGlobant\G4G\2026\4950- [RAS]- For due date contracts block revenue actions\GLB951-66593-GLB951-66595 - Create a New Project with a Waiver\Test_Cases_GLB951-66593_GLB951-66595.xlsx'

output_lines = []

try:
    wb = openpyxl.load_workbook(path)
    output_lines.append(f'Sheets: {wb.sheetnames}')
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        output_lines.append(f'\n=== Sheet: {sheet} ===')
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 100:
                output_lines.append(f'Row {i+1}: {row}')
except Exception as e:
    output_lines.append(f'Error: {e}')

with open(r'C:\Users\luis.acosta\Gearrdo-Personal\Tesis\excel_output.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))

print('Done! Output written to excel_output.txt')
